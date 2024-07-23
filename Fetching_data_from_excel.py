import openpyxl

wb = openpyxl.load_workbook("datadriventests.xlsx")

#cur_sheet = wb.get_sheet_by_name("Data")
cur_sheet = wb.active

row = cur_sheet.max_row
cols = cur_sheet.max_column

for r in range(1, row+1):
    for c in range(1, cols+1):
        print(cur_sheet.cell(row=r, column=c).value, end="   ")

    print("")